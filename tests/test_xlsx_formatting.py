"""Tests for XLSX conditional formatting."""

import pytest
from openpyxl import Workbook, load_workbook
from pathlib import Path

from xlsx_tools.formatting import apply_excel_formatting


@pytest.fixture
def sample_xlsx(tmp_path):
    path = str(tmp_path / "data.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Product", "Revenue", "Margin"])
    ws.append(["A", 5000, 0.15])
    ws.append(["B", 12000, 0.35])
    ws.append(["C", 8000, 0.22])
    ws.append(["D", 3000, 0.08])
    wb.save(path)
    return path


class TestHighlightRule:
    def test_greater_than(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [{"range": "B2:B5", "type": "highlight", "condition": ">", "value": 10000, "color": "#00FF00"}]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "1 formatting rule" in result
        wb = load_workbook(out)
        assert len(wb["Sales"].conditional_formatting) == 1

    def test_less_than(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [{"range": "B2:B5", "type": "highlight", "condition": "<", "value": 5000, "color": "FF0000"}]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "1" in result

    def test_between(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [{"range": "B2:B5", "type": "highlight", "condition": "between", "value": [5000, 10000], "color": "FFFF00"}]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "1" in result


class TestColorScale:
    def test_color_scale(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [{"range": "B2:B5", "type": "color_scale", "min_color": "#FF0000", "max_color": "#00FF00"}]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "1" in result


class TestDataBar:
    def test_data_bar(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [{"range": "B2:B5", "type": "data_bar", "color": "638EC6"}]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "1" in result


class TestMultipleRules:
    def test_multiple_rules(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        rules = [
            {"range": "B2:B5", "type": "color_scale", "min_color": "FF0000", "max_color": "00FF00"},
            {"range": "C2:C5", "type": "highlight", "condition": ">", "value": 0.2, "color": "00FF00"},
            {"range": "C2:C5", "type": "data_bar", "color": "4472C4"},
        ]
        result = apply_excel_formatting(sample_xlsx, "Sales", rules, out)
        assert "3 formatting rule" in result


class TestErrors:
    def test_invalid_sheet(self, sample_xlsx):
        with pytest.raises(ValueError, match="not found"):
            apply_excel_formatting(sample_xlsx, "NonExistent", [])
