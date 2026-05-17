"""Tests for Excel chart generation."""

import pytest
from openpyxl import Workbook, load_workbook
from pathlib import Path

from xlsx_tools.charts import create_excel_chart


@pytest.fixture
def sample_xlsx(tmp_path):
    path = str(tmp_path / "data.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Month", "Revenue", "Profit"])
    ws.append(["Jan", 10000, 3000])
    ws.append(["Feb", 12000, 4000])
    ws.append(["Mar", 15000, 5500])
    ws.append(["Apr", 11000, 3200])
    wb.save(path)
    return path


class TestCreateChart:
    def test_bar_chart_from_file(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        result = create_excel_chart(
            file_path=sample_xlsx, chart_type="bar", title="Revenue",
            sheet_name="Sales", data_start_column=2, data_end_column=3, output_path=out,
        )
        assert "bar" in result
        assert Path(out).exists()
        wb = load_workbook(out)
        assert len(wb.active._charts) == 1

    def test_line_chart(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        result = create_excel_chart(
            file_path=sample_xlsx, chart_type="line", title="Trend",
            sheet_name="Sales", data_start_column=2, output_path=out,
        )
        assert "line" in result

    def test_pie_chart(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        result = create_excel_chart(
            file_path=sample_xlsx, chart_type="pie", title="Distribution",
            sheet_name="Sales", data_start_column=2, output_path=out,
        )
        assert "pie" in result

    def test_column_chart(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        result = create_excel_chart(
            file_path=sample_xlsx, chart_type="column", title="Comparison",
            sheet_name="Sales", data_start_column=2, data_end_column=3, output_path=out,
        )
        assert "column" in result

    def test_inline_data(self, tmp_path):
        out = str(tmp_path / "new.xlsx")
        data = [
            ["Product", "Sales"],
            ["Widget A", 500],
            ["Widget B", 300],
            ["Widget C", 800],
        ]
        result = create_excel_chart(
            chart_type="bar", title="Product Sales",
            inline_data=data, output_path=out,
        )
        assert Path(out).exists()
        wb = load_workbook(out)
        assert wb.active.max_row == 4

    def test_invalid_chart_type(self, sample_xlsx):
        with pytest.raises(ValueError, match="Unsupported"):
            create_excel_chart(file_path=sample_xlsx, chart_type="scatter", title="X")

    def test_invalid_sheet(self, sample_xlsx):
        with pytest.raises(ValueError, match="not found"):
            create_excel_chart(file_path=sample_xlsx, chart_type="bar", title="X", sheet_name="NonExistent")

    def test_multi_series(self, sample_xlsx, tmp_path):
        out = str(tmp_path / "out.xlsx")
        result = create_excel_chart(
            file_path=sample_xlsx, chart_type="bar", title="Multi",
            sheet_name="Sales", data_start_column=2, data_end_column=3, output_path=out,
        )
        assert "Multi" in result
        wb = load_workbook(out)
        assert len(wb.active._charts) == 1
