"""Edit existing XLSX files — modify cells, insert/delete rows."""

from openpyxl import load_workbook


def edit_xlsx_cell(file_path: str, sheet_name: str, cell: str, value: str, output_path: str | None = None) -> str:
    """Edit a specific cell value. Cell format: 'A1', 'B2', etc."""
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    ws[cell] = value

    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()
    return f"Cell {cell} in '{sheet_name}' set to '{value}'. Saved to {save_path}"


def insert_xlsx_row(file_path: str, sheet_name: str, row_index: int, values: list[str], output_path: str | None = None) -> str:
    """Insert a new row at the given index (1-based)."""
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    ws.insert_rows(row_index)
    for col, val in enumerate(values, 1):
        ws.cell(row=row_index, column=col, value=val)

    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()
    return f"Row inserted at index {row_index} in '{sheet_name}'. Saved to {save_path}"


def delete_xlsx_row(file_path: str, sheet_name: str, row_index: int, output_path: str | None = None) -> str:
    """Delete a row at the given index (1-based)."""
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    if row_index < 1 or row_index > ws.max_row:
        raise ValueError(f"Row index {row_index} out of range (1-{ws.max_row})")

    ws.delete_rows(row_index)

    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()
    return f"Row {row_index} deleted from '{sheet_name}'. Saved to {save_path}"
