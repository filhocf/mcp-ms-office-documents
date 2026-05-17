"""XLSX conditional formatting — rules-based cell highlighting."""

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def apply_excel_formatting(file_path: str, sheet_name: str, rules: list[dict], output_path: str | None = None) -> str:
    """Apply conditional formatting rules to an Excel sheet.

    Each rule dict:
    - range: cell range (e.g. "B2:B100")
    - type: "highlight", "color_scale", "data_bar"
    - For highlight: condition (">", "<", "between", "equal"), value, color
    - For color_scale: min_color, max_color
    - For data_bar: color
    """
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    applied = 0

    for rule in rules:
        cell_range = rule.get("range")
        rule_type = rule.get("type")

        if not cell_range or not rule_type:
            continue

        if rule_type == "highlight":
            condition = rule.get("condition", ">")
            value = rule.get("value", 0)
            color = rule.get("color", "FFFF00").lstrip("#")

            operator_map = {
                ">": "greaterThan",
                "<": "lessThan",
                ">=": "greaterThanOrEqual",
                "<=": "lessThanOrEqual",
                "==": "equal",
                "equal": "equal",
                "between": "between",
            }
            op = operator_map.get(condition, "greaterThan")

            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            if op == "between" and isinstance(value, list) and len(value) == 2:
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator=op, formula=[str(value[0]), str(value[1])], fill=fill)
                )
            else:
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator=op, formula=[str(value)], fill=fill)
                )
            applied += 1

        elif rule_type == "color_scale":
            min_color = rule.get("min_color", "FF0000").lstrip("#")
            max_color = rule.get("max_color", "00FF00").lstrip("#")

            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min", start_color=min_color,
                    end_type="max", end_color=max_color,
                )
            )
            applied += 1

        elif rule_type == "data_bar":
            color = rule.get("color", "638EC6").lstrip("#")
            ws.conditional_formatting.add(
                cell_range,
                DataBarRule(start_type="min", end_type="max", color=color)
            )
            applied += 1

    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()
    return f"Applied {applied} formatting rule(s) to '{sheet_name}'. Saved to {save_path}"
