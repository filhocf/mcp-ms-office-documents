"""Excel chart builder using openpyxl charts."""

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter


CHART_TYPES = {
    "bar": BarChart,
    "column": BarChart,
    "line": LineChart,
    "pie": PieChart,
}


def create_excel_chart(
    file_path: str | None = None,
    chart_type: str = "bar",
    title: str = "Chart",
    sheet_name: str = "Sheet1",
    data_range: str | None = None,
    categories_column: int = 1,
    data_start_column: int = 2,
    data_end_column: int | None = None,
    start_row: int = 1,
    end_row: int | None = None,
    chart_position: str = "E2",
    inline_data: list[list] | None = None,
    output_path: str | None = None,
    file_name: str | None = None,
) -> str:
    """Create a chart in an Excel file.

    Can either:
    - Add chart to existing file (file_path) using data already in a sheet
    - Create new file with inline_data and chart

    Args:
        file_path: Path to existing XLSX (or None to create new)
        chart_type: bar, column, line, pie
        title: Chart title
        sheet_name: Sheet containing data
        categories_column: Column number for categories (1-based)
        data_start_column: First data column (1-based)
        data_end_column: Last data column (None = same as start)
        start_row: First row of data (1 = header)
        end_row: Last row (None = auto-detect)
        chart_position: Cell where chart is placed (e.g. "E2")
        inline_data: Data to write before charting [[header...], [row1...], ...]
        output_path: Where to save (None = overwrite or auto-generate)
        file_name: Name for upload (when no output_path)
    """
    if chart_type not in CHART_TYPES:
        raise ValueError(f"Unsupported chart type '{chart_type}'. Supported: {list(CHART_TYPES.keys())}")

    # Load or create workbook
    if file_path:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write inline data if provided
    if inline_data:
        for row_data in inline_data:
            ws.append(row_data)

    # Auto-detect end_row
    if end_row is None:
        end_row = ws.max_row
    if data_end_column is None:
        data_end_column = data_start_column

    # Create chart
    ChartClass = CHART_TYPES[chart_type]
    chart = ChartClass()
    chart.title = title

    if chart_type == "column":
        chart.type = "col"

    # Data reference
    data_ref = Reference(ws, min_col=data_start_column, min_row=start_row, max_col=data_end_column, max_row=end_row)
    cats_ref = Reference(ws, min_col=categories_column, min_row=start_row + 1, max_row=end_row)

    if chart_type == "pie":
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
    else:
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

    ws.add_chart(chart, chart_position)

    # Save
    save_path = output_path or file_path
    if save_path:
        wb.save(save_path)
        wb.close()
        return f"Chart '{title}' ({chart_type}) added to {save_path}"
    else:
        # Upload via upload_tools
        import io
        from upload_tools.main import upload_file
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()
        result = upload_file(buffer, "xlsx", file_name=file_name)
        return result
